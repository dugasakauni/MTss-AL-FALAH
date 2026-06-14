from flask import Flask, render_template, request, redirect, flash, send_file, session
import sqlite3
import os
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table,
    TableStyle
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

app = Flask(__name__)
app.secret_key = "Alfalah2026"

# =========================
# DATABASE
# =========================

# Pastikan folder instance ada
os.makedirs('instance', exist_ok=True)

def get_db():
    conn = sqlite3.connect('instance/pendaftaran.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()

    conn.execute("""
    CREATE TABLE IF NOT EXISTS pendaftaran (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nomor_pendaftaran TEXT NOT NULL,
        nama TEXT NOT NULL,
        tempat_lahir TEXT NOT NULL,
        tanggal_lahir TEXT NOT NULL,
        jenis_kelamin TEXT NOT NULL,
        asal_sekolah TEXT NOT NULL,
        nama_ortu TEXT NOT NULL,
        no_hp TEXT NOT NULL,
        alamat TEXT NOT NULL
    )
    """)

    conn.commit()
    conn.close()

# Jalankan saat aplikasi pertama kali start
init_db()

# =========================
# HALAMAN UTAMA
# =========================

@app.route('/')
def index():
    return render_template('index.html')

# =========================
# SIMPAN PENDAFTARAN
# =========================

@app.route('/daftar', methods=['POST'])
def daftar():

    nama = request.form['nama']
    tempat_lahir = request.form['tempat_lahir']
    tanggal_lahir = request.form['tanggal_lahir']
    jenis_kelamin = request.form['jenis_kelamin']
    asal_sekolah = request.form['asal_sekolah']
    nama_ortu = request.form['nama_ortu']
    no_hp = request.form['no_hp']
    alamat = request.form['alamat']

    conn = get_db()
    jumlah = conn.execute("""
    SELECT COUNT(*) as total
    FROM pendaftaran
    """).fetchone()

    nomor_urut = jumlah["total"] + 1

    nomor_pendaftaran = f"PPDB-2026-{nomor_urut:04d}"

    conn.execute("""
        INSERT INTO pendaftaran (
            nomor_pendaftaran,
            nama,
            tempat_lahir,
            tanggal_lahir,
            jenis_kelamin,
            asal_sekolah,
            nama_ortu,
            no_hp,
            alamat
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        nomor_pendaftaran,
        nama,
        tempat_lahir,
        tanggal_lahir,
        jenis_kelamin,
        asal_sekolah,
        nama_ortu,
        no_hp,
        alamat
    ))

    conn.commit()
    conn.close()

    return render_template(
        'sukses.html',
        nomor_pendaftaran=nomor_pendaftaran,
        nama=nama
    )

@app.route('/bukti/<nomor>')
def bukti_pdf(nomor):

    conn = get_db()

    siswa = conn.execute("""
        SELECT *
        FROM pendaftaran
        WHERE nomor_pendaftaran = ?
    """, (nomor,)).fetchone()

    conn.close()

    if not siswa:
        return "Data tidak ditemukan"

    filename = os.path.join(
        app.root_path,
        f"bukti_{nomor}.pdf"
    )

    doc = SimpleDocTemplate(filename)

    styles = getSampleStyleSheet()
    from reportlab.lib.enums import TA_CENTER

    styles['Heading2'].alignment = TA_CENTER
    styles['Title'].alignment = TA_CENTER

    content = []

    logo_path = os.path.join(
        app.root_path,
        "static",
        "logo.jpg"
    )

    logo = Image(
        logo_path,
        width=70,
        height=70
    )
    content.append(logo)
    content.append(Spacer(1, 10))

    content.append(
        Paragraph(
            "BUKTI PENDAFTARAN PPDB",
            styles['Title']
        )
    )
    content.append(
        Paragraph(
            "MTss Al-Falah Penjaringan",
            styles['Title']
        )
    )

    content.append(
        Paragraph(
            "Tahun Pelajaran 2026/2027",
            styles['Heading2']
        )
    )

    content.append(Spacer(1, 20))

    data = [
        ["Nomor Pendaftaran", siswa['nomor_pendaftaran']],
        ["Nama Lengkap", siswa['nama']],
        ["Tempat Lahir", siswa['tempat_lahir']],
        ["Tanggal Lahir", siswa['tanggal_lahir']],
        ["Jenis Kelamin", siswa['jenis_kelamin']],
        ["Asal Sekolah", siswa['asal_sekolah']],
        ["Nama Orang Tua", siswa['nama_ortu']],
        ["No HP", siswa['no_hp']],
        ["Alamat", siswa['alamat']]
    ]

    table = Table(
        data,
        colWidths=[150, 300]
    )

    table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))

    content.append(table)

    content.append(Spacer(1, 20))

    content.append(
        Paragraph(
            "<b>Status:</b> Pendaftaran Berhasil Diterima",
            styles['Normal']
        )
    )

    doc.build(content)

    return send_file(
        filename,
        as_attachment=True
    )
# =========================
# HALAMAN ADMIN
# =========================

@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect('/login')

    conn = get_db()

    semua_siswa = conn.execute("""
        SELECT *
        FROM pendaftaran
        ORDER BY id DESC
    """).fetchall()

    conn.close()

    return render_template(
        'admin.html',
        semua_siswa=semua_siswa
    )

@app.route('/hapus/<int:id>')
def hapus(id):

    conn = get_db()

    conn.execute(
        "DELETE FROM pendaftaran WHERE id=?",
        (id,)
    )

    conn.commit()
    conn.close()

    return redirect('/admin')

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        username = request.form['username']
        password = request.form['password']

        if username == 'operator' and password == 'Alfalah2026':

            session['admin'] = True
            return redirect('/admin')

        flash('Username atau Password salah!')
        return redirect('/login')

    return render_template('login.html')

@app.route('/logout')
def logout():

    session.clear()

    return redirect('/login')

# =========================
# EXPORT EXCEL
# =========================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border, Side
from datetime import datetime

@app.route('/export')
def export():

    conn = get_db()

    data = conn.execute("""
        SELECT *
        FROM pendaftaran
        ORDER BY id DESC
    """).fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Pendaftar"
    ws.merge_cells('A1:I1')
    ws['A1'] = "DATA PENDAFTAR PPDB MTs AL-FALAH"

    ws['A1'].font = Font(
        bold=True,
        size=16
    )

    ws['A1'].alignment = Alignment(
        horizontal='center'
    )

    ws.merge_cells('A2:I2')
    ws['A2'] = "Tahun Pelajaran 2026/2027"

    ws['A2'].alignment = Alignment(
        horizontal='center'
    )
    ws.merge_cells('A3:I3')
    ws['A3'] = f"Dicetak pada: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws['A3'].alignment = Alignment(horizontal='center')

    headers = [
        "No",
        "Nama Lengkap",
        "Tempat Lahir",
        "Tanggal Lahir",
        "Jenis Kelamin",
        "Asal Sekolah",
        "Nama Orang Tua",
        "No HP",
        "Alamat"
    ]

    # Header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="6B7280",
            end_color="6B7280",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Isi Data
    row_excel = 5
    for no, row in enumerate(data, start=1):
        ws.cell(row=row_excel, column=1, value=no)
        ws.cell(row=row_excel, column=2, value=row["nama"])
        ws.cell(row=row_excel, column=3, value=row["tempat_lahir"])
        ws.cell(row=row_excel, column=4, value=row["tanggal_lahir"])
        ws.cell(row=row_excel, column=5, value=row["jenis_kelamin"])
        ws.cell(row=row_excel, column=6, value=row["asal_sekolah"])
        ws.cell(row=row_excel, column=7, value=row["nama_ortu"])
        ws.cell(row=row_excel, column=8, value=row["no_hp"])
        ws.cell(row=row_excel, column=9, value=row["alamat"])

        row_excel += 1

    # Atur lebar kolom
    column_widths = {
        "A": 8,
        "B": 25,
        "C": 20,
        "D": 18,
        "E": 15,
        "F": 25,
        "G": 25,
        "H": 18,
        "I": 40
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    file_name = os.path.join(
        app.root_path,
        "data_pendaftar.xlsx"
    )
    thin = Side(
        border_style="thin",
        color="000000"
    )

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row,
        min_col=1,
        max_col=9
    ):
        for cell in row:
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )
    wb.save(file_name)

    return send_file(
        file_name,
        as_attachment=True
    )

# =========================
# RUN APP
# =========================

if __name__ == '__main__':
    app.run()